"""SimpleClaw Admin — FastAPI router 工厂。

用法（在 server.py startup 中）：
    from admin.routes import make_admin_router
    app.include_router(make_admin_router(...))
"""

from __future__ import annotations

import asyncio
import io
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from fastapi import APIRouter, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, Response
from loguru import logger
from starlette.requests import Request

from admin.prompt_files import make_prompt_file_map

if TYPE_CHECKING:
    from simpleclaw.llm.base import LLMProvider
    from Mojing.storage.document_repo import DocumentRepository
    from Mojing.storage.memory_repo import MySQLMemory
    from Mojing.storage.runtime_task_repo import RuntimeTaskRepository
    from Mojing.storage.session_repo import SessionRepository
    from Mojing.storage.tenant_state_repo import TenantStateRepository
    from Mojing.storage.database import Database


_PREVIEW_DIVIDER = "═" * 60


def _assemble_runtime_prompt(
    *,
    stable_prefix: str,
    dynamic_parts: list[str],
    reminder: str,
    tool_schemas: list[dict],
) -> str:
    """把运行时实际下发到 LLM 的内容拼成可读文本（用于 admin 预览，仅展示用）。

    布局：
      ── STATIC（system 主体，享 prefix cache）
      ── DYNAMIC（每轮变 —— USER.md / 时间 / 自拍时距 等，附在 system 末尾）
      ── SYSTEM ACTIVATION / REMINDER（如有）
      ── TOOLS（registry.schemas() —— 通过 tools= 字段下发给 LLM）
    """
    lines: list[str] = []

    lines.append(_PREVIEW_DIVIDER)
    lines.append("STATIC PROMPT  ·  stable_prefix（享 prefix cache）")
    lines.append(_PREVIEW_DIVIDER)
    lines.append(stable_prefix or "(空)")

    lines.append("")
    lines.append(_PREVIEW_DIVIDER)
    lines.append(f"DYNAMIC SECTIONS  ·  {len(dynamic_parts)} 段（拼到 system 末尾）")
    lines.append(_PREVIEW_DIVIDER)
    if dynamic_parts:
        lines.append("\n\n---\n\n".join(p.strip() for p in dynamic_parts if p))
    else:
        lines.append("(无)")

    lines.append("")
    lines.append(_PREVIEW_DIVIDER)
    lines.append("SYSTEM ACTIVATION / REMINDER")
    lines.append(_PREVIEW_DIVIDER)
    lines.append(reminder.strip() if reminder else "(无)")

    lines.append("")
    lines.append(_PREVIEW_DIVIDER)
    lines.append(f"TOOLS  ·  {len(tool_schemas)} 个（通过 tools= 字段下发）")
    lines.append(_PREVIEW_DIVIDER)
    if tool_schemas:
        import json as _json
        lines.append(_json.dumps(tool_schemas, ensure_ascii=False, indent=2))
    else:
        lines.append("(无)")

    return "\n".join(lines)


def _scenario_from_admin_payload(body: dict) -> dict:
    """Convert Scenario Lab JSON into the runner's scenario shape."""
    if not isinstance(body, dict):
        raise ValueError("payload must be a JSON object")
    agent = str(body.get("agent") or "main").strip()
    if agent not in {"main", "skin_diary", "deep_report"}:
        raise ValueError(f"unsupported agent: {agent!r}")
    stage = str(body.get("initial_stage") or body.get("stage") or "novice").strip() or "novice"
    raw_turns = body.get("turns") or []
    if not isinstance(raw_turns, list):
        raise ValueError("turns must be a list")
    turns = _admin_turns_from_raw(raw_turns, max_turns=200)
    if not turns:
        raise ValueError("至少需要一轮 query 或图片 URL")

    scenario_id = str(body.get("id") or f"admin_scenario_{datetime.now().strftime('%Y%m%d%H%M%S')}").strip()
    scenario = {
        "id": scenario_id,
        "name": str(body.get("name") or scenario_id),
        "description": str(body.get("description") or "Admin Scenario Lab run"),
        "agent": agent,
        "initial_stage": stage,
        "wait_side_effects_s": float(body.get("wait_side_effects_s") or 0),
        "turns": turns,
    }
    tenant_key = str(body.get("tenant_key") or "").strip()
    if tenant_key:
        scenario["tenant_key"] = tenant_key
    seed = body.get("seed")
    if isinstance(seed, dict) and seed:
        scenario["seed"] = seed
    scenario.update(_admin_replay_entry_from_payload(body))
    return scenario


def _admin_replay_entry_from_payload(body: dict) -> dict[str, Any]:
    raw_surface = str(body.get("replay_surface") or body.get("entry") or "app_chat").strip().lower()
    if raw_surface in {"v1", "v1_device", "device", "hardware", "hardware_v1"}:
        device_id = str(body.get("device_id") or body.get("deviceId") or "").strip()
        device_code = str(body.get("device_code") or body.get("deviceCode") or "").strip()
        custom = dict(body.get("custom") or {}) if isinstance(body.get("custom"), dict) else {}
        if device_id:
            custom["device_id"] = device_id
        if device_code:
            custom["device_code"] = device_code
        return {
            "replay_surface": "v1_device",
            "protocol": "v1_chat_completions",
            "endpoint": "/v1/chat/completions",
            "prompt_surface": "device",
            "device_id": device_id,
            "device_code": device_code,
            "custom": custom,
        }
    return {
        "replay_surface": "app_chat",
        "protocol": "agent_chat",
        "endpoint": "/agent/chat",
        "prompt_surface": "app",
    }


def _admin_turns_from_raw(raw_turns: object, *, max_turns: int) -> list[dict]:
    if not isinstance(raw_turns, list):
        raise ValueError("turns must be a list")
    if len(raw_turns) > max_turns:
        raise ValueError(f"最多支持 {max_turns} 轮")

    turns: list[dict] = []
    for item in raw_turns:
        if isinstance(item, str):
            user = item.strip()
            media: list[str] = []
        elif isinstance(item, dict):
            user = str(item.get("user") or item.get("query") or "").strip()
            media = _normalize_media(item.get("media") or item.get("images") or [])
        else:
            continue
        if not user and not media:
            continue
        turn_no = len(turns) + 1
        payload = {
            "user": user,
            "media": media,
            "phase": "measured",
            "measure": True,
            "capture_prompt": True,
            "capture_prompt_chars": 1_000_000 if turn_no == 1 else 120_000,
            "hard_assertions": [{"reply_non_empty": True}],
        }
        if isinstance(item, dict):
            wait_after = item.get("wait_after_s", item.get("wait_side_effects_s", item.get("delay_s")))
            if wait_after not in (None, ""):
                payload["wait_side_effects_s"] = float(wait_after)
        turns.append(payload)
    return turns


def _scenario_suite_from_admin_payload(body: dict) -> dict:
    if not isinstance(body, dict):
        raise ValueError("payload must be a JSON object")
    stage = str(body.get("initial_stage") or body.get("stage") or "novice").strip() or "novice"
    suite_id = str(body.get("id") or f"admin_suite_{datetime.now().strftime('%Y%m%d%H%M%S')}").strip()
    tenant_key = str(body.get("tenant_key") or _short_admin_tenant_key()).strip()
    wait_s = float(body.get("wait_side_effects_s") or 0)
    windows_raw = body.get("windows") or {}
    if not isinstance(windows_raw, dict):
        raise ValueError("windows must be an object")

    windows: dict[str, dict] = {}
    total_turns = 0
    for agent in ("main", "skin_diary", "deep_report"):
        raw = windows_raw.get(agent) or {}
        if isinstance(raw, list):
            raw_turns = raw
        elif isinstance(raw, dict):
            raw_turns = raw.get("turns") or []
        else:
            raw_turns = []
        turns = _admin_turns_from_raw(raw_turns, max_turns=200) if raw_turns else []
        total_turns += len(turns)
        windows[agent] = {
            "agent": agent,
            "session_key": str(
                (raw.get("session_key") if isinstance(raw, dict) else "")
                or _default_admin_session_key(agent, tenant_key)
            ),
            "turns": turns,
        }
    if total_turns <= 0:
        raise ValueError("至少需要在一个窗口填写一轮 query 或图片 URL")

    seed = body.get("seed")
    return {
        "id": suite_id,
        "name": str(body.get("name") or suite_id),
        "tenant_key": tenant_key,
        "initial_stage": stage,
        "wait_side_effects_s": wait_s,
        "seed": seed if isinstance(seed, dict) else {},
        "replay": _admin_replay_entry_from_payload(body),
        "windows": windows,
    }


def _default_admin_session_key(agent: str, tenant_key: str) -> str:
    if agent == "main":
        return f"main:{tenant_key}"
    if agent == "skin_diary":
        return f"skin_diary:{tenant_key}"
    if agent == "deep_report":
        return f"deep_report:{tenant_key}"
    return f"{agent}:{tenant_key}"


def _short_admin_tenant_key() -> str:
    """Use short test IDs so external image/report services share exact IDs."""
    return f"test_{datetime.now().strftime('%H%M%S%f')[:12]}"


def _normalize_media(raw: object) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, str):
        parts = raw.replace(",", "\n").splitlines()
        return [p.strip() for p in parts if p.strip()]
    if isinstance(raw, list):
        return [str(item).strip() for item in raw if str(item or "").strip()]
    return []


_URL_RE = re.compile(r"https?://[^\s,，;；)）\"'\\\\]+")
_TIME_TEXT_RE = re.compile(r"^\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}")
_TIME_HEADERS = {
    "发生时间", "时间", "发送时间", "消息时间", "创建时间", "create_time",
    "created_at", "createdat", "time", "timestamp",
}
_USER_HEADERS = {
    "user query", "user_query", "query", "用户消息", "用户输入", "用户发送消息",
    "用户发送的消息", "用户说", "用户query", "用户 query", "用户提问",
    "消息内容", "文本内容", "内容", "消息", "发送内容", "对话内容",
    "问题", "text", "content", "body",
}
_MEDIA_HEADERS = {
    "图片", "图片url", "图片 url", "image", "image_url", "image url",
    "media", "media_url", "url", "图片链接", "图片地址", "附件", "素材",
    "voice_url", "voice url",
}
_WAIT_HEADERS = {"等待秒", "等待时间", "间隔秒", "间隔时间", "delay", "delay_s", "wait", "wait_s"}
_ROLE_HEADERS = {"角色", "发送方", "消息角色", "role", "from", "sender"}
_RAW_HEADERS = {"raw", "原始数据", "payload"}
_MESSAGE_TYPE_HEADERS = {"message_type", "messagetype", "消息类型", "类型"}
_VOICE_TEXT_HEADERS = {"voice_text", "voice text", "语音文字", "语音文本"}
_ASSISTANT_HEADERS = {
    "agent输出", "agent output", "agent回复", "assistant", "assistant_reply",
    "assistant reply", "模型回复", "回复", "原始agent回复", "产品测试agent回复",
}
_TOOL_NAME_HEADERS = {"tool_name", "tool name", "工具名", "工具"}
_REVIEW_MODE_HEADERS = {
    "标注类型", "标注", "评分标注", "评分类型", "评分", "review_mode",
    "review mode", "mode",
}
_PRODUCT_FEEDBACK_HEADERS = {"产品反馈", "反馈", "问题反馈", "改进原因", "feedback", "product_feedback"}
_USER_ROLE_VALUES = {"user", "human", "用户", "客户", "用户消息", "用户发送", "真人用户"}
_ASSISTANT_ROLE_VALUES = {
    "assistant", "agent", "ai", "bot", "机器人", "助手", "主agent",
    "主 agent", "agent回复", "assistant回复",
}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\n", "").replace("：", ":")


def _compact_header(value: str) -> str:
    return re.sub(r"[\s_\-:：/]+", "", value.lower())


def _find_header_cols(headers: list[str], exact: set[str], contains: tuple[str, ...] = ()) -> list[int]:
    exact_compact = {_compact_header(item) for item in exact}
    cols: list[int] = []
    for idx, header in enumerate(headers):
        compact = _compact_header(header)
        if header in exact or compact in exact_compact:
            cols.append(idx)
            continue
        if any(token in header or _compact_header(token) in compact for token in contains):
            cols.append(idx)
    return cols


def _extract_urls(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in _URL_RE.findall(text) if item.strip()]


def _strip_urls(value: Any) -> str:
    return _URL_RE.sub("", _stringify_cell(value)).strip(" \t\r\n,，;；")


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def _looks_like_time_cell(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    text = _stringify_cell(value)
    return bool(_TIME_TEXT_RE.search(text))


def _parse_wait_seconds(value: Any) -> float:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return max(float(value), 0)
    text = _stringify_cell(value)
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0
    return max(float(match.group(0)), 0)


def _loads_jsonish(value: Any) -> Any:
    text = _stringify_cell(value)
    if not text or text[:1] not in "{[":
        return None
    try:
        return json.loads(text)
    except Exception:
        return None


def _text_from_message_content(content: Any) -> str:
    if isinstance(content, str):
        text = _strip_urls(content)
        return "" if not text or not re.search(r"[\u4e00-\u9fffA-Za-z]", text) else text
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if not isinstance(item, dict):
                continue
            if item.get("type") == "text" or "text" in item:
                text = _text_from_message_content(item.get("text"))
                if text:
                    parts.append(text)
        return "\n".join(parts).strip()
    return ""


def _is_persisted_session_user_row(values: list[Any], raw_cols: list[int], message_type_cols: list[int]) -> bool:
    has_empty_message_type = bool(message_type_cols) and all(
        idx >= len(values) or not _stringify_cell(values[idx])
        for idx in message_type_cols
    )
    if not has_empty_message_type:
        return False
    for idx in raw_cols:
        if idx >= len(values):
            continue
        payload = _loads_jsonish(values[idx])
        if isinstance(payload, dict) and payload.get("session_key") and payload.get("seq") is not None:
            return True
    return False


def _role_kind(value: Any) -> str:
    text = _normalize_header(value)
    compact = _compact_header(text)
    if not compact:
        return ""
    user_values = {_compact_header(item) for item in _USER_ROLE_VALUES}
    assistant_values = {_compact_header(item) for item in _ASSISTANT_ROLE_VALUES}
    if compact in user_values or any(token in compact for token in ("user", "human", "用户", "客户")):
        return "user"
    if compact in assistant_values or any(token in compact for token in ("assistant", "agent", "bot", "助手", "机器人")):
        return "assistant"
    return ""


def _is_role_label(value: Any) -> bool:
    compact = _compact_header(_normalize_header(value))
    if not compact:
        return False
    role_values = {_compact_header(item) for item in (_USER_ROLE_VALUES | _ASSISTANT_ROLE_VALUES)}
    return compact in role_values


def _best_fallback_user_cell(values: list[Any], headers: list[str]) -> str:
    best_score = 0
    best_text = ""
    noisy_headers = {
        "tenant", "tenantkey", "tenantid", "userid", "session", "sessionkey",
        "sessionid", "id", "序号", "轮次",
    }
    for idx, value in enumerate(values):
        header = headers[idx] if idx < len(headers) else ""
        compact_header = _compact_header(header)
        text = _strip_urls(value)
        if not text or _looks_like_time_cell(value) or _is_role_label(value):
            continue
        if compact_header in {
            _compact_header(item)
            for item in (
                _TIME_HEADERS | _MEDIA_HEADERS | _WAIT_HEADERS | _ROLE_HEADERS
                | _MESSAGE_TYPE_HEADERS | _VOICE_TEXT_HEADERS | _ASSISTANT_HEADERS
                | _REVIEW_MODE_HEADERS | _PRODUCT_FEEDBACK_HEADERS
            )
        }:
            continue
        if compact_header in noisy_headers or compact_header in {_compact_header(item) for item in _RAW_HEADERS}:
            continue
        score = min(len(text), 200)
        if any(token in header for token in ("消息", "内容", "query", "文本", "问题", "提问", "发送")):
            score += 80
        if re.search(r"[\u4e00-\u9fffA-Za-z]", text):
            score += 20
        if _extract_urls(value):
            score -= 20
        if score > best_score:
            best_score = score
            best_text = text
    return best_text


def _strip_chat_prefix(text: str) -> str:
    value = str(text or "").strip()
    if value.startswith("[CHAT]"):
        value = value[len("[CHAT]"):].strip()
    return value


def _normalize_review_mode(value: Any) -> str:
    text = _normalize_header(value)
    compact = _compact_header(text)
    if not compact:
        return "stable"
    if compact in {"提升", "需要提升", "改进", "需要改进", "improve", "fix", "bad", "不好"}:
        return "improve"
    if compact in {"保持稳定", "稳定", "stable", "keep", "good", "正常", "好"}:
        return "stable"
    if compact in {"不计分", "跳过", "不评分", "skip", "none", "ignore", "观察"}:
        return "none"
    return "stable"


def _row_role(values: list[Any], role_cols: list[int]) -> str:
    for idx in role_cols:
        if idx < len(values):
            role = _role_kind(values[idx])
            if role:
                return role
    return ""


def _row_text(values: list[Any], cols: list[int]) -> str:
    for idx in cols:
        if idx < len(values):
            text = _strip_chat_prefix(_stringify_cell(values[idx]))
            if text:
                return text
    return ""


def _attach_xlsx_baseline_replies(
    turns: list[dict[str, Any]],
    *,
    rows: list[Any],
    headers: list[str],
    role_cols: list[int],
    user_cols: list[int],
    tool_name_cols: list[int],
    review_mode_cols: list[int],
    product_feedback_cols: list[int],
    header_index: int,
) -> None:
    if not turns:
        return
    turn_rows = [int(item.get("row") or 0) for item in turns]
    row_by_no = {
        row_no: list(row)
        for row_no, row in enumerate(rows[header_index + 1 :], start=header_index + 2)
    }
    for index, turn in enumerate(turns):
        if str(turn.get("baseline_reply") or "").strip():
            continue
        start = turn_rows[index]
        end = turn_rows[index + 1] if index + 1 < len(turn_rows) else len(rows) + 1
        normal_parts: list[str] = []
        chat_parts: list[str] = []
        annotated_review_mode = ""
        annotated_product_feedback = ""
        for row_no in range(start + 1, end):
            values = row_by_no.get(row_no)
            if not values:
                continue
            if _row_role(values, role_cols) != "assistant":
                continue
            if any(_stringify_cell(values[idx]) for idx in tool_name_cols if idx < len(values)):
                continue
            raw_text = _row_text(values, user_cols)
            if not raw_text:
                continue
            first_content = _stringify_cell(values[user_cols[0]]) if user_cols and user_cols[0] < len(values) else ""
            row_review_mode = ""
            for idx in review_mode_cols:
                if idx < len(values) and _stringify_cell(values[idx]):
                    row_review_mode = _normalize_review_mode(values[idx])
                    break
            row_product_feedback = ""
            for idx in product_feedback_cols:
                if idx < len(values):
                    row_product_feedback = _stringify_cell(values[idx])
                    if row_product_feedback:
                        break
            if first_content.startswith("[CHAT]"):
                chat_parts.append(raw_text)
                if row_review_mode or row_product_feedback:
                    annotated_review_mode = row_review_mode or annotated_review_mode
                    annotated_product_feedback = row_product_feedback or annotated_product_feedback
            else:
                normal_parts.append(raw_text)
                if (row_review_mode or row_product_feedback) and not annotated_review_mode and not annotated_product_feedback:
                    annotated_review_mode = row_review_mode
                    annotated_product_feedback = row_product_feedback
        if chat_parts:
            turn["baseline_reply"] = chat_parts[-1]
        elif normal_parts:
            turn["baseline_reply"] = "\n".join(dict.fromkeys(normal_parts))
        if annotated_review_mode:
            turn["review_mode"] = annotated_review_mode
        if annotated_product_feedback:
            turn["product_feedback"] = annotated_product_feedback


def _parse_xlsx_turns(file_bytes: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，无法解析 xlsx") from exc

    # Some exported product-test workbooks carry stale sheet dimension metadata.
    # openpyxl read_only mode trusts that metadata and may only expose A1.
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"sheet": sheet.title, "turns": [], "headers": [], "row_count": 0, "raw_rows": [], "skipped_rows": []}

    header_index = 0
    headers = [_normalize_header(v) for v in rows[0]]
    for idx, row in enumerate(rows[:10]):
        normalized = [_normalize_header(v) for v in row]
        if (
            _find_header_cols(normalized, _USER_HEADERS, ("消息", "内容", "query", "文本", "问题", "提问"))
            or _find_header_cols(normalized, _TIME_HEADERS, ("时间", "time"))
            or _find_header_cols(normalized, _MEDIA_HEADERS, ("图片", "image", "url", "附件"))
            or _find_header_cols(normalized, _WAIT_HEADERS, ("等待", "间隔", "delay", "wait"))
        ):
            header_index = idx
            headers = normalized
            break

    time_cols = _find_header_cols(headers, _TIME_HEADERS, ("时间", "time"))
    user_cols = _find_header_cols(headers, _USER_HEADERS, ("消息", "内容", "query", "文本", "问题", "提问"))
    media_cols = _find_header_cols(headers, _MEDIA_HEADERS, ("图片", "image", "url", "附件", "素材"))
    wait_cols = _find_header_cols(headers, _WAIT_HEADERS, ("等待", "间隔", "delay", "wait"))
    role_cols = _find_header_cols(headers, _ROLE_HEADERS, ("角色", "发送方", "role", "sender"))
    raw_cols = _find_header_cols(headers, _RAW_HEADERS, ("raw",))
    message_type_cols = _find_header_cols(headers, _MESSAGE_TYPE_HEADERS, ("message_type", "消息类型"))
    voice_text_cols = _find_header_cols(headers, _VOICE_TEXT_HEADERS, ("voice_text", "语音"))
    assistant_cols = _find_header_cols(headers, _ASSISTANT_HEADERS, ("agent回复", "assistant_reply"))
    tool_name_cols = _find_header_cols(headers, _TOOL_NAME_HEADERS, ("tool_name", "工具名"))
    review_mode_cols = _find_header_cols(headers, _REVIEW_MODE_HEADERS, ("标注", "评分类型", "review_mode"))
    product_feedback_cols = _find_header_cols(headers, _PRODUCT_FEEDBACK_HEADERS, ("产品反馈", "问题反馈", "feedback"))

    turns: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    for row_no, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = list(row)
        if len(raw_rows) < 30 and any(_stringify_cell(v) for v in values):
            raw_rows.append({"row": row_no, "values": [_stringify_cell(v) for v in values]})

        role = ""
        role_text = ""
        for idx in role_cols:
            if idx < len(values):
                role_text = _stringify_cell(values[idx])
                role = _role_kind(values[idx])
                if role:
                    break
        if role_cols and role != "user":
            if len(skipped_rows) < 30 and any(_stringify_cell(v) for v in values):
                reason = "assistant_row" if role == "assistant" else f"non_user_role:{role_text or '-'}"
                skipped_rows.append({"row": row_no, "reason": reason, "values": [_stringify_cell(v) for v in values]})
            continue
        if _is_persisted_session_user_row(values, raw_cols, message_type_cols):
            if len(skipped_rows) < 30:
                skipped_rows.append({"row": row_no, "reason": "persisted_duplicate_user_row", "values": [_stringify_cell(v) for v in values]})
            continue
        if role == "assistant":
            if len(skipped_rows) < 30:
                skipped_rows.append({"row": row_no, "reason": "assistant_row", "values": [_stringify_cell(v) for v in values]})
            continue

        user = ""
        for idx in user_cols:
            if idx < len(values):
                candidate = _strip_urls(values[idx])
                if candidate:
                    user = candidate
                    break
        if not user:
            for idx in voice_text_cols:
                if idx < len(values):
                    candidate = _strip_urls(values[idx])
                    if candidate:
                        user = candidate
                        break
        if not user:
            user = _best_fallback_user_cell(values, headers)

        media: list[str] = []
        for idx in media_cols:
            if idx < len(values):
                media.extend(_extract_urls(values[idx]))
        for value in values:
            media.extend(_extract_urls(value))
        media = list(dict.fromkeys(media))

        occurred_at = ""
        for idx in time_cols:
            if idx < len(values):
                occurred_at = _stringify_cell(values[idx])
                if occurred_at:
                    break

        wait_after_s = 0.0
        for idx in wait_cols:
            if idx < len(values):
                wait_after_s = _parse_wait_seconds(values[idx])
                if wait_after_s:
                    break

        review_mode = "stable"
        for idx in review_mode_cols:
            if idx < len(values):
                review_mode = _normalize_review_mode(values[idx])
                break

        product_feedback = ""
        for idx in product_feedback_cols:
            if idx < len(values):
                product_feedback = _stringify_cell(values[idx])
                if product_feedback:
                    break

        if not user and not media:
            if len(skipped_rows) < 30 and any(_stringify_cell(v) for v in values):
                skipped_rows.append({"row": row_no, "reason": "empty_or_unrecognized", "values": [_stringify_cell(v) for v in values]})
            continue
        baseline_reply = ""
        for idx in assistant_cols:
            if idx < len(values):
                baseline_reply = _strip_chat_prefix(_stringify_cell(values[idx]))
                if baseline_reply:
                    break
        turns.append({
            "row": row_no,
            "occurred_at": occurred_at,
            "user": user,
            "media": media,
            "wait_after_s": wait_after_s,
            "baseline_reply": baseline_reply,
            "review_mode": review_mode,
            "product_feedback": product_feedback,
        })

    _attach_xlsx_baseline_replies(
        turns,
        rows=rows,
        headers=headers,
        role_cols=role_cols,
        user_cols=user_cols,
        tool_name_cols=tool_name_cols,
        review_mode_cols=review_mode_cols,
        product_feedback_cols=product_feedback_cols,
        header_index=header_index,
    )

    review_counts: dict[str, int] = {"stable": 0, "improve": 0, "none": 0}
    feedback_count = 0
    score_ready_count = 0
    for turn in turns:
        mode = str(turn.get("review_mode") or "stable")
        review_counts[mode] = review_counts.get(mode, 0) + 1
        if str(turn.get("product_feedback") or "").strip():
            feedback_count += 1
        if mode == "improve" and str(turn.get("product_feedback") or "").strip():
            score_ready_count += 1

    return {
        "sheet": sheet.title,
        "headers": headers,
        "header_row": header_index + 1,
        "row_count": max(len(rows) - header_index - 1, 0),
        "detected": {
            "time_cols": time_cols,
            "user_cols": user_cols,
            "media_cols": media_cols,
            "wait_cols": wait_cols,
            "role_cols": role_cols,
            "raw_cols": raw_cols,
            "message_type_cols": message_type_cols,
            "voice_text_cols": voice_text_cols,
            "assistant_cols": assistant_cols,
            "review_mode_cols": review_mode_cols,
            "product_feedback_cols": product_feedback_cols,
        },
        "raw_rows": raw_rows,
        "skipped_rows": skipped_rows,
        "review_counts": review_counts,
        "feedback_count": feedback_count,
        "score_ready_count": score_ready_count,
        "turns": turns,
    }


def _admin_source_session_key(value: str) -> str:
    session = str(value or "").strip()
    if session and ":" not in session:
        return f"main:{session}"
    return session


def _jsonish(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value)
    if not text:
        return ""
    if text[:1] in "{[":
        try:
            return json.loads(text)
        except Exception:
            return text
    return text


def _short_text(value: Any, limit: int = 500) -> str:
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(_jsonish(value), ensure_ascii=False)
    return text if len(text) <= limit else text[:limit] + "..."


def _project_env(name: str, default: str = "") -> str:
    value = os.getenv(name, "").strip()
    if value:
        return value
    env_path = Path(__file__).resolve().parents[1] / ".env"
    try:
        for line in env_path.read_text(encoding="utf-8").splitlines():
            raw = line.strip()
            if not raw or raw.startswith("#") or "=" not in raw:
                continue
            key, raw_value = raw.split("=", 1)
            if key.strip() == name:
                return raw_value.strip().strip('"').strip("'")
    except FileNotFoundError:
        pass
    except Exception:
        logger.debug("failed to read project .env for {}", name)
    return default


def _extract_json_object(text: str) -> dict[str, Any]:
    value = str(text or "").strip()
    if value.startswith("```"):
        value = re.sub(r"^```(?:json)?\s*", "", value)
        value = re.sub(r"\s*```$", "", value)
    try:
        parsed = json.loads(value)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        pass
    start = value.find("{")
    end = value.rfind("}")
    if start >= 0 and end > start:
        parsed = json.loads(value[start:end + 1])
        if isinstance(parsed, dict):
            return parsed
    raise ValueError("模型没有返回合法 JSON")


def _stability_card_prompt(payload: dict[str, Any]) -> list[dict[str, str]]:
    mode = str(payload.get("mode") or "stable").strip()
    weight = payload.get("weight") or 1.0
    user = str(payload.get("user") or "").strip()
    media = payload.get("media") or []
    baseline_reply = str(payload.get("baseline_reply") or "").strip()
    product_feedback = str(payload.get("product_feedback") or "").strip()
    tools = payload.get("tools") or []
    system = (
        "你是产品回归测试评分卡生成器。你不修改 Agent 回复，只把人工确认或产品反馈"
        "转成后续可用于 LLM Judge 的结构化评分卡。必须只输出 JSON，不要输出 Markdown。"
    )
    user_prompt = {
        "task": "generate_replay_scoring_card",
        "mode": mode,
        "weight": weight,
        "input": {
            "user_query": user,
            "media": media,
            "baseline_reply": baseline_reply,
            "product_feedback": product_feedback,
            "baseline_tools": tools,
        },
        "requirements": [
            "stable 模式：baseline_reply 是产品认可或认为正常的原始回复；生成用于判断候选回复是否保持稳定的评分卡。",
            "improve 模式：baseline_reply 是当时不满意的原始回复，product_feedback 是产品指出的问题；生成用于判断候选回复是否修复问题的评分卡。",
            "不要要求逐字一致；关注核心意图、承诺边界、关键动作、用户体验质量。",
            "hard_checks 只写可由代码或日志确定的检查；judge_instruction 写给后续 LLM Judge。",
            "输出字段固定：mode, weight, baseline_summary, problem_summary, desired_behavior, hard_checks, risk_guardrails, judge_instruction, scoring, notes。",
        ],
        "output_schema": {
            "mode": "stable|improve",
            "weight": "number",
            "baseline_summary": "string",
            "problem_summary": "string",
            "desired_behavior": "string",
            "hard_checks": ["string"],
            "risk_guardrails": ["string"],
            "judge_instruction": "string",
            "scoring": {
                "pass": "string",
                "warn": "string",
                "fail": "string",
            },
            "notes": "string",
        },
    }
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False)},
    ]


async def _generate_stability_card(payload: dict[str, Any]) -> dict[str, Any]:
    provider = _project_env("GEPA_REFLECTION_PROVIDER", "kimi").lower()
    if provider != "kimi":
        raise RuntimeError(f"当前仅支持 GEPA_REFLECTION_PROVIDER=kimi，实际为 {provider!r}")
    api_key = _project_env("KIMI_API_KEY") or _project_env("MOONSHOT_API_KEY")
    if not api_key:
        raise RuntimeError("缺少 KIMI_API_KEY（或 MOONSHOT_API_KEY）")
    api_base = _project_env("KIMI_API_BASE", "https://api.moonshot.cn/v1").rstrip("/")
    model = _project_env("KIMI_MODEL", "kimi-k2.6")
    timeout_s = float(_project_env("KIMI_TIMEOUT_S", "60") or 60)
    max_tokens = int(_project_env("KIMI_MAX_TOKENS", "1600") or 1600)
    temperature = float(_project_env("KIMI_TEMPERATURE", "0.6") or 0.6)
    from script.gepa.reflection import KimiStreamingReflectionLM

    lm = KimiStreamingReflectionLM(
        model,
        Path("script/gepa/runs/admin_stability_cards"),
        api_key=api_key,
        api_base=api_base,
        max_tokens=max_tokens,
        timeout=timeout_s,
        thinking_type=_project_env("KIMI_THINKING_TYPE", "disabled") or "disabled",
        include_usage=True,
        temperature=temperature,
    )
    content = await asyncio.to_thread(lm, _stability_card_prompt(payload))
    card = _extract_json_object(content)
    card["generated_by"] = {"provider": "kimi", "model": model}
    return card


def _kimi_reflection_lm(run_name: str):
    provider = _project_env("GEPA_REFLECTION_PROVIDER", "kimi").lower()
    if provider != "kimi":
        raise RuntimeError(f"当前仅支持 GEPA_REFLECTION_PROVIDER=kimi，实际为 {provider!r}")
    api_key = _project_env("KIMI_API_KEY") or _project_env("MOONSHOT_API_KEY")
    if not api_key:
        raise RuntimeError("缺少 KIMI_API_KEY（或 MOONSHOT_API_KEY）")
    api_base = _project_env("KIMI_API_BASE", "https://api.moonshot.cn/v1").rstrip("/")
    model = _project_env("KIMI_MODEL", "kimi-k2.6")
    timeout_s = float(_project_env("KIMI_TIMEOUT_S", "60") or 60)
    max_tokens = int(_project_env("KIMI_MAX_TOKENS", "1600") or 1600)
    temperature = float(_project_env("KIMI_TEMPERATURE", "0.6") or 0.6)
    from script.gepa.reflection import KimiStreamingReflectionLM

    return KimiStreamingReflectionLM(
        model,
        Path("script/gepa/runs") / run_name,
        api_key=api_key,
        api_base=api_base,
        max_tokens=max_tokens,
        timeout=timeout_s,
        thinking_type=_project_env("KIMI_THINKING_TYPE", "disabled") or "disabled",
        include_usage=True,
        temperature=temperature,
    ), model


def _scenario_check_prompt(payload: dict[str, Any]) -> list[dict[str, str]]:
    system = (
        "你是产品回归测试 LLM Judge。你只根据给定的 user_query、baseline_reply、"
        "product_feedback、score_card、candidate_reply 和关键日志给单轮回归测试打分。"
        "必须只输出 JSON，不要输出 Markdown。不要因为候选回复和 baseline 不逐字一致就扣分。"
    )
    output_schema = {
        "turn": "number",
        "mode": "stable|improve",
        "score": "0到1的小数",
        "verdict": "pass|warn|fail",
        "reason": "string",
        "strengths": ["string"],
        "issues": ["string"],
        "evidence": ["string"],
        "suggested_fix": "string",
    }
    user_prompt = {
        "task": "judge_replay_turn",
        "rubric": [
            "stable：判断 candidate_reply 是否保持 baseline_reply 的核心意图、关键动作、承诺边界和用户体验；允许自然改写。",
            "improve：判断 candidate_reply 是否修复 product_feedback 指出的问题，同时保留 baseline 中合理部分，且没有新增错误或过度承诺。",
            "如果回复为空、明显答非所问、反向违背反馈，score 应低于 0.4。",
            "如果基本达标但有轻微话术或信息遗漏，score 在 0.7 到 0.89。",
            "如果完全达标，score 在 0.9 到 1.0。",
            "工具调用和 provider/event 日志只作为证据，不要求每轮都必须调用工具；除非 score_card 明确要求。",
        ],
        "input": payload,
        "output_schema": output_schema,
    }
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False)},
    ]


def _admin_report_measured_turns(report: dict[str, Any]) -> list[dict[str, Any]]:
    main = ((report or {}).get("windows") or {}).get("main") or {}
    turns = main.get("turns") or []
    if not isinstance(turns, list):
        return []
    return [
        turn for turn in turns
        if isinstance(turn, dict) and (not turn.get("phase") or turn.get("phase") == "measured" or turn.get("measure") is True)
    ]


def _parse_score_card(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {"raw": value.strip()}
    return {}


def _turn_check_context(turn: dict[str, Any]) -> dict[str, Any]:
    tools = []
    for item in turn.get("tools") or []:
        if not isinstance(item, dict):
            continue
        tools.append({
            "tool_name": item.get("tool_name"),
            "status": item.get("status") or item.get("action") or ("failed" if item.get("ok") is False else "ok"),
            "duration_ms": item.get("duration_ms"),
            "arguments": item.get("arguments"),
            "result_preview": item.get("result_preview") or _short_text(item.get("result"), 500),
        })
    provider_sources = []
    for capture in (turn.get("attention") or [])[:3]:
        if not isinstance(capture, dict):
            continue
        for packet in (capture.get("packets") or [])[:8]:
            if isinstance(packet, dict):
                provider_sources.append({
                    "source": packet.get("source") or packet.get("provider"),
                    "placement": packet.get("placement"),
                    "content_preview": packet.get("content_preview"),
                })
    pushes = []
    delta = ((turn.get("state_changes") or {}).get("session_delta") or {})
    for msg in (delta.get("new_messages") or [])[:10]:
        if isinstance(msg, dict) and msg.get("role") != "user":
            pushes.append({"role": msg.get("role"), "content": msg.get("content"), "seq": msg.get("seq")})
    return {
        "tools": tools[:8],
        "provider": provider_sources[:12],
        "event_push": pushes,
        "runtime_tasks": (turn.get("runtime_tasks_created") or [])[:8],
        "business_jobs": turn.get("business_jobs") or {},
    }


async def _check_scenario_report(payload: dict[str, Any]) -> dict[str, Any]:
    report = payload.get("report") or {}
    annotations = payload.get("annotations") or []
    if not isinstance(report, dict):
        raise ValueError("report must be a JSON object")
    if not isinstance(annotations, list):
        raise ValueError("annotations must be a list")
    measured = _admin_report_measured_turns(report)
    if not measured:
        raise ValueError("没有 measured turn，请先运行回放")

    annotation_by_turn: dict[int, dict[str, Any]] = {}
    for idx, item in enumerate(annotations, start=1):
        if not isinstance(item, dict):
            continue
        turn_no = int(item.get("turn") or idx)
        annotation_by_turn[turn_no] = item

    lm = None
    model = _project_env("KIMI_MODEL", "kimi-k2.6")
    results: list[dict[str, Any]] = []
    for idx, turn in enumerate(measured, start=1):
        turn_no = int(turn.get("turn") or idx)
        annotation = annotation_by_turn.get(turn_no) or annotation_by_turn.get(idx) or {}
        mode = str(annotation.get("review_mode") or "stable").strip()
        if mode == "none":
            results.append({
                "turn": turn_no,
                "mode": "none",
                "weight": 0,
                "score": None,
                "weighted_score": None,
                "verdict": "skip",
                "reason": "不计分",
            })
            continue
        baseline_reply = str(annotation.get("baseline_reply") or "").strip()
        if not baseline_reply:
            results.append({
                "turn": turn_no,
                "mode": mode,
                "weight": 0,
                "score": None,
                "weighted_score": None,
                "verdict": "skip",
                "reason": "缺少 baseline_reply，无法判断稳定性或提升效果",
            })
            continue
        weight = float(annotation.get("weight") or 1.0)
        check_payload = {
            "turn": turn_no,
            "mode": mode,
            "weight": weight,
            "user_query": str(turn.get("user") or annotation.get("user") or "").strip(),
            "media": turn.get("media") or annotation.get("media") or [],
            "baseline_reply": baseline_reply,
            "product_feedback": str(annotation.get("product_feedback") or "").strip(),
            "score_card": _parse_score_card(annotation.get("score_card")),
            "candidate_reply": str(turn.get("reply_preview") or "").strip(),
            "timing": turn.get("timing") or {},
            "logs": _turn_check_context(turn),
        }
        if lm is None:
            lm, model = _kimi_reflection_lm("admin_replay_checks")
        try:
            content = await asyncio.to_thread(lm, _scenario_check_prompt(check_payload))
            judged = _extract_json_object(content)
        except Exception as exc:
            logger.exception("admin scenario check turn failed turn={}", turn_no)
            results.append({
                "turn": turn_no,
                "mode": mode,
                "weight": weight,
                "score": 0.0,
                "weighted_score": 0.0,
                "verdict": "fail",
                "reason": f"评分失败：{exc}",
                "issues": [str(exc)],
                "evidence": [],
                "generated_by": {"provider": "kimi", "model": model},
            })
            continue
        score = max(0.0, min(float(judged.get("score") or 0), 1.0))
        verdict = str(judged.get("verdict") or "").lower()
        if verdict not in {"pass", "warn", "fail"}:
            verdict = "pass" if score >= 0.9 else "warn" if score >= 0.7 else "fail"
        judged.update({
            "turn": turn_no,
            "mode": mode,
            "weight": weight,
            "score": score,
            "weighted_score": score * weight,
            "verdict": verdict,
            "generated_by": {"provider": "kimi", "model": model},
        })
        results.append(judged)

    scored = [r for r in results if isinstance(r.get("score"), (int, float))]
    total_weight = sum(float(r.get("weight") or 0) for r in scored)
    weighted_score = sum(float(r.get("weighted_score") or 0) for r in scored)
    stable_scores = [float(r["score"]) for r in scored if r.get("mode") == "stable"]
    improve_scores = [float(r["score"]) for r in scored if r.get("mode") == "improve"]
    low_turns = [
        {"turn": r.get("turn"), "mode": r.get("mode"), "score": r.get("score"), "verdict": r.get("verdict"), "reason": r.get("reason")}
        for r in scored
        if float(r.get("score") or 0) < 0.8
    ]
    return {
        "summary": {
            "total_turns": len(measured),
            "scored_turns": len(scored),
            "skipped_turns": len(results) - len(scored),
            "weighted_average": (weighted_score / total_weight) if total_weight else None,
            "plain_average": (sum(float(r["score"]) for r in scored) / len(scored)) if scored else None,
            "stable_average": (sum(stable_scores) / len(stable_scores)) if stable_scores else None,
            "improve_average": (sum(improve_scores) / len(improve_scores)) if improve_scores else None,
            "low_turns": low_turns,
        },
        "results": results,
    }


async def _snapshot_preview(db, *, tenant_key: str, session_key: str, snapshot_at: str) -> dict[str, Any]:
    from script.runner.runner import _normalize_snapshot_at, _resolve_snapshot_cutoff

    tenant_key = str(tenant_key or "").strip()
    session_key = _admin_source_session_key(session_key)
    snapshot_at = _normalize_snapshot_at(snapshot_at)
    if not tenant_key or not session_key or not snapshot_at:
        raise ValueError("tenant_key、session_key、snapshot_at 都不能为空")

    cutoff = await _resolve_snapshot_cutoff(
        db,
        src_tenant=tenant_key,
        src_session=session_key,
        snapshot_at=snapshot_at,
        inclusive=False,
    )

    async with db.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT tenant_key, session_key, session_type, origin_session_key,
                       channel, title, is_primary, last_consolidated, created_at, updated_at
                FROM nb_sessions
                WHERE tenant_key=%s AND session_key=%s
                LIMIT 1
                """,
                (tenant_key, session_key),
            )
            session_row = await cur.fetchone()

            await cur.execute(
                """
                SELECT seq, role, tool_name, message_json, created_at
                FROM nb_session_messages
                WHERE tenant_key=%s AND session_key=%s AND seq<=%s
                ORDER BY seq ASC
                """,
                (tenant_key, session_key, cutoff),
            )
            message_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT doc_type, doc_name, content, version_no, change_source, created_at
                FROM nb_tenant_document_versions
                WHERE tenant_key=%s AND created_at<=%s
                ORDER BY doc_type, doc_name, version_no DESC
                """,
                (tenant_key, snapshot_at),
            )
            doc_version_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT source, memory_type, topic, description, content, created_at, updated_at
                FROM nb_memory_entries
                WHERE tenant_key=%s AND created_at<=%s
                ORDER BY created_at ASC
                LIMIT 200
                """,
                (tenant_key, snapshot_at),
            )
            memory_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT profile_id, session_key, image_url, analysis_id, overall_state,
                       skin_attribute_json, advantages_json, signals_json, sync_status, created_at
                FROM nb_tenant_skin_profiles
                WHERE tenant_key=%s AND created_at<=%s
                ORDER BY created_at DESC
                LIMIT 10
                """,
                (tenant_key, snapshot_at),
            )
            profile_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT id, state, summary, chips, analyzed_at, create_time
                FROM nb_skin_diary_results
                WHERE tenant_key=%s AND deleted=0 AND create_time<=%s
                ORDER BY create_time DESC
                LIMIT 10
                """,
                (tenant_key, snapshot_at),
            )
            diary_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT id, brand, product_name, category, in_cabinet,
                       usage_status, commercial_image, user_photo, create_time
                FROM nb_skincare_cabinet_product
                WHERE user_id=%s AND deleted=0 AND create_time<=%s
                ORDER BY create_time ASC
                LIMIT 200
                """,
                (tenant_key, snapshot_at),
            )
            cabinet_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT report_id, session_id, status, summary, create_time
                FROM nb_slow_model_reports
                WHERE user_id=%s AND deleted=0 AND status='done' AND create_time<=%s
                ORDER BY create_time DESC
                LIMIT 5
                """,
                (tenant_key, snapshot_at),
            )
            report_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT job_id, session_key, image_ref, focus, status, summary_text, created_at
                FROM nb_image_analysis_jobs
                WHERE tenant_key=%s AND status IN ('succeeded','user_md_synced') AND created_at<=%s
                ORDER BY created_at DESC
                LIMIT 10
                """,
                (tenant_key, snapshot_at),
            )
            image_rows = await cur.fetchall()

            await cur.execute(
                """
                SELECT journey_json, primary_session_key, last_user_activity_at, updated_at
                FROM nb_tenant_state
                WHERE tenant_key=%s
                LIMIT 1
                """,
                (tenant_key,),
            )
            state_row = await cur.fetchone()

    docs: dict[tuple[Any, Any], Any] = {}
    for row in doc_version_rows:
        key = (row[0], row[1])
        if key not in docs:
            docs[key] = row

    messages = []
    for row in message_rows[-120:]:
        parsed = _jsonish(row[3])
        content = parsed.get("content") if isinstance(parsed, dict) else ""
        messages.append({
            "seq": row[0],
            "role": row[1],
            "tool_name": row[2],
            "created_at": str(row[4]) if row[4] else "",
            "content": _short_text(content, 700),
        })

    return {
        "source": {
            "tenant_key": tenant_key,
            "session_key": session_key,
            "snapshot_at": snapshot_at,
            "msg_seq_cutoff": cutoff,
        },
        "session": {
            "exists": bool(session_row),
            "row": {
                "tenant_key": session_row[0],
                "session_key": session_row[1],
                "session_type": session_row[2],
                "origin_session_key": session_row[3],
                "channel": session_row[4],
                "title": session_row[5],
                "is_primary": session_row[6],
                "last_consolidated": session_row[7],
                "created_at": str(session_row[8]) if session_row[8] else "",
                "updated_at": str(session_row[9]) if session_row[9] else "",
            } if session_row else None,
        },
        "messages": {
            "count_before_cutoff": len(message_rows),
            "showing_last": len(messages),
            "items": messages,
        },
        "documents": [
            {
                "doc_type": row[0],
                "doc_name": row[1],
                "version_no": row[3],
                "change_source": row[4],
                "created_at": str(row[5]) if row[5] else "",
                "content": str(row[2] or ""),
            }
            for row in docs.values()
        ],
        "memories": [
            {
                "source": row[0],
                "memory_type": row[1],
                "topic": row[2],
                "description": row[3],
                "content": row[4],
                "created_at": str(row[5]) if row[5] else "",
                "updated_at": str(row[6]) if row[6] else "",
            }
            for row in memory_rows
        ],
        "skin_profiles": [
            {
                "profile_id": row[0],
                "session_key": row[1],
                "image_url": row[2],
                "analysis_id": row[3],
                "overall_state": row[4],
                "skin_attribute": _jsonish(row[5]),
                "advantages": _jsonish(row[6]),
                "signals": _jsonish(row[7]),
                "sync_status": row[8],
                "created_at": str(row[9]) if row[9] else "",
            }
            for row in profile_rows
        ],
        "skin_diaries": [
            {
                "id": row[0],
                "state": row[1],
                "summary": row[2],
                "chips": _jsonish(row[3]),
                "analyzed_at": str(row[4]) if row[4] else "",
                "create_time": str(row[5]) if row[5] else "",
            }
            for row in diary_rows
        ],
        "cabinet_products": [
            {
                "id": row[0],
                "brand": row[1],
                "product_name": row[2],
                "category": row[3],
                "in_cabinet": row[4],
                "usage_status": row[5],
                "commercial_image": row[6],
                "user_photo": row[7],
                "create_time": str(row[8]) if row[8] else "",
            }
            for row in cabinet_rows
        ],
        "deep_reports": [
            {
                "report_id": row[0],
                "session_id": row[1],
                "status": row[2],
                "summary": row[3],
                "create_time": str(row[4]) if row[4] else "",
            }
            for row in report_rows
        ],
        "image_jobs": [
            {
                "job_id": row[0],
                "session_key": row[1],
                "image_ref": row[2],
                "focus": row[3],
                "status": row[4],
                "summary_text": row[5],
                "created_at": str(row[6]) if row[6] else "",
            }
            for row in image_rows
        ],
        "tenant_state": {
            "journey": _jsonish(state_row[0]) if state_row else None,
            "primary_session_key": state_row[1] if state_row else "",
            "last_user_activity_at": str(state_row[2]) if state_row and state_row[2] else "",
            "updated_at": str(state_row[3]) if state_row and state_row[3] else "",
        },
    }


def make_admin_router(
    workspace: Path,
    subagent_prompt: Path,
    *,
    db=None,
    document_repo=None,
    runtime_task_repo=None,
    session_repo=None,
    tenant_state_repo=None,
    llm=None,
    main_agent=None,
    skin_diary_subagent=None,
    deep_report_subagent=None,
) -> APIRouter:
    """返回注册了所有 /admin/* 路由的 APIRouter。

    Args:
        workspace:            Mojing/workspace/ 目录
        subagent_prompt:      Mojing/subagent/prompt/ 目录
        db:                   Database 实例（用于 MySQLMemory）
        document_repo:        DocumentRepository（USER.md 读写）
        runtime_task_repo:    RuntimeTaskRepository（后台任务状态）
        session_repo:         SessionRepository（历史消息）
        tenant_state_repo:    TenantStateRepository（Journey stage）
        llm:                  LLMProvider（冷链路 / postprocess LLM 调用）
        main_agent:           MainAgent 实例（preview 用 make_tool_registry）
        skin_diary_subagent:  SkinDiarySubagent 实例（同上，肌肤日记视角）
        deep_report_subagent: DeepReportSubagent 实例（同上，深度报告视角）
    """
    _prompt_map = make_prompt_file_map(workspace, subagent_prompt)
    router = APIRouter(prefix="/admin", tags=["admin"])

    # ------------------------------------------------------------------
    # 懒加载 HTML 页面（避免循环 import）
    # ------------------------------------------------------------------

    _page_cache: list[str] = []

    def _get_page() -> str:
        if not _page_cache:
            from admin._page import HTML_PAGE
            _page_cache.append(HTML_PAGE)
        return _page_cache[0]

    # ================================================================
    # Prompt 文件管理
    # ================================================================

    @router.get("/editor", response_class=HTMLResponse)
    async def admin_editor():
        """返回 Admin UI 页面。"""
        return HTMLResponse(_get_page())

    @router.get("/scenario", response_class=HTMLResponse)
    async def scenario_lab():
        """返回多轮对话测试台页面。"""
        from admin.scenario_page import SCENARIO_HTML_PAGE
        return HTMLResponse(SCENARIO_HTML_PAGE)

    @router.post("/scenario/import-xlsx")
    async def import_scenario_xlsx(file: UploadFile = File(...)):
        """解析产品测试 Excel，把用户消息和图片 URL 转成 Scenario turns。"""
        filename = str(file.filename or "")
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            return JSONResponse({"error": "请上传 .xlsx / .xlsm 文件"}, status_code=400)
        try:
            payload = _parse_xlsx_turns(await file.read())
        except Exception as exc:
            logger.exception("admin scenario import xlsx failed")
            return JSONResponse({"error": str(exc)}, status_code=500)
        return JSONResponse(payload)

    @router.post("/scenario/stability-card")
    async def scenario_stability_card(request: Request):
        """调用 Kimi 生成单轮稳定/改进评分卡，不执行评分。"""
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": f"bad request: {exc}"}, status_code=400)
        try:
            card = await _generate_stability_card(body if isinstance(body, dict) else {})
        except Exception as exc:
            logger.exception("admin scenario stability card failed")
            return JSONResponse({"error": str(exc)}, status_code=500)
        return JSONResponse({"card": card})

    @router.post("/scenario/check")
    async def scenario_check(request: Request):
        """对已完成的回放结果执行 stable/improve 回归评分。"""
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": f"bad request: {exc}"}, status_code=400)
        try:
            payload = await _check_scenario_report(body if isinstance(body, dict) else {})
        except Exception as exc:
            logger.exception("admin scenario check failed")
            return JSONResponse({"error": str(exc)}, status_code=500)
        return JSONResponse(payload)

    @router.post("/scenario/snapshot-preview")
    async def scenario_snapshot_preview(request: Request):
        """只读预览指定 tenant/session/时间点会被快照克隆的数据库状态。"""
        if db is None:
            return JSONResponse({"error": "db not configured"}, status_code=503)
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": f"bad request: {exc}"}, status_code=400)
        try:
            payload = await _snapshot_preview(
                db,
                tenant_key=str(body.get("tenant_key") or body.get("tenant") or "").strip(),
                session_key=str(body.get("session_key") or body.get("session") or "").strip(),
                snapshot_at=str(body.get("snapshot_at") or body.get("snapshotAt") or "").strip(),
            )
        except Exception as exc:
            logger.exception("admin scenario snapshot preview failed")
            return JSONResponse({"error": str(exc)}, status_code=500)
        return JSONResponse(payload)

    @router.get("/prompt")
    async def get_prompt(file: str = ""):
        """读取 prompt 文件内容。"""
        entry = _prompt_map.get(file)
        if entry is None:
            return JSONResponse({"error": f"unknown file key: {file!r}"}, status_code=404)
        try:
            content = entry.path.read_text(encoding="utf-8") if entry.path.exists() else ""
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=500)
        return JSONResponse({
            "content":    content,
            "hot_reload": entry.hot_reload,
            "label":      entry.label,
            "path":       str(entry.path),
        })

    @router.put("/prompt")
    async def put_prompt(request: Request, file: str = ""):
        """保存 prompt 文件内容到磁盘。"""
        entry = _prompt_map.get(file)
        if entry is None:
            return Response(f"unknown file key: {file!r}", status_code=404)
        try:
            body = await request.body()
            content = body.decode("utf-8")
        except Exception as exc:
            return Response(f"bad request: {exc}", status_code=400)
        try:
            entry.path.parent.mkdir(parents=True, exist_ok=True)
            entry.path.write_text(content, encoding="utf-8")
            logger.info("admin: wrote {} ({} bytes)", entry.path, len(content))
        except Exception as exc:
            return Response(str(exc), status_code=500)
        return Response(status_code=204)

    @router.post("/restart")
    async def restart():
        """平滑重启服务（新进程启动后再自杀）。"""
        async def _do_restart():
            await asyncio.sleep(0.4)
            cmd = [sys.executable] + sys.argv
            subprocess.Popen(cmd, start_new_session=True,
                             stdout=open("/tmp/simpleclaw_restart.log", "w"),
                             stderr=subprocess.STDOUT)
            await asyncio.sleep(1.5)
            import os, signal
            os.kill(os.getpid(), signal.SIGTERM)

        asyncio.create_task(_do_restart())
        return Response(status_code=202)

    # ================================================================
    # Scenario Lab
    # ================================================================

    @router.post("/scenario/run")
    async def run_scenario_lab(request: Request):
        """运行前端提交的多轮对话场景，返回完整观测报告。"""
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": f"bad request: {exc}"}, status_code=400)

        try:
            scenario = _scenario_from_admin_payload(body)
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

        try:
            from script.runner.runner import run_scenario_dict
            container = getattr(request.app.state, "container", None)
            run_dir = Path("script/logs") / f"admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            result = await run_scenario_dict(
                scenario,
                run_dir=run_dir,
                source="admin/scenario",
                container=container,
            )
            return JSONResponse(result)
        except Exception as exc:
            logger.exception("admin scenario run failed")
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.post("/scenario/suite/run")
    async def run_scenario_suite_lab(request: Request):
        """运行三窗口 Agent 调试场景，返回主/子 Agent 分窗观测报告。"""
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": f"bad request: {exc}"}, status_code=400)

        try:
            suite = _scenario_suite_from_admin_payload(body)
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

        try:
            from script.runner.runner import run_scenario_dict, _session_messages
            container = getattr(request.app.state, "container", None)
            run_dir = Path("script/logs") / f"admin_suite_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            windows: dict[str, dict] = {}
            seed_consumed = False
            for agent in ("main", "skin_diary", "deep_report"):
                window = suite["windows"][agent]
                turns = window.get("turns") or []
                if not turns:
                    windows[agent] = {
                        "agent": agent,
                        "session_key": window["session_key"],
                        "turns": [],
                        "verdict": "SKIPPED",
                    }
                    continue
                scenario = {
                    "id": f"{suite['id']}_{agent}",
                    "name": f"{suite['name']} / {agent}",
                    "description": "Admin Scenario Suite window run",
                    "agent": agent,
                    "tenant_key": suite["tenant_key"],
                    "session_key": window["session_key"],
                    "initial_stage": suite["initial_stage"],
                    "wait_side_effects_s": suite["wait_side_effects_s"],
                    "turns": turns,
                }
                scenario.update(suite.get("replay") or {})
                if not seed_consumed and suite.get("seed"):
                    scenario["seed"] = suite["seed"]
                    seed_consumed = True
                result = await run_scenario_dict(
                    scenario,
                    run_dir=run_dir,
                    source="admin/scenario-suite",
                    container=container,
                )
                windows[agent] = result

            sessions = {}
            if container is not None:
                for agent in ("main", "skin_diary", "deep_report"):
                    session_key = suite["windows"][agent]["session_key"]
                    sessions[agent] = await _session_messages(
                        container,
                        suite["tenant_key"],
                        session_key,
                    )

            verdicts = [
                item.get("verdict")
                for item in windows.values()
                if item.get("verdict") != "SKIPPED"
            ]
            return JSONResponse({
                "report_format": "scenario_suite_v1",
                "scenario": suite["id"],
                "tenant_key": suite["tenant_key"],
                "initial_stage": suite["initial_stage"],
                "replay": suite.get("replay") or {},
                "verdict": "PASS" if verdicts and all(v == "PASS" for v in verdicts) else "FAIL",
                "windows": windows,
                "sessions": sessions,
                "run_dir": str(run_dir),
            })
        except Exception as exc:
            logger.exception("admin scenario suite run failed")
            return JSONResponse({"error": str(exc)}, status_code=500)

    # ================================================================
    # 租户数据
    # ================================================================

    @router.get("/tenants")
    async def list_tenants():
        """列出所有有历史消息记录的租户 key。"""
        if session_repo is None:
            return JSONResponse({"tenants": []})
        try:
            async with db.acquire() as conn:
                async with conn.cursor() as cur:
                    await cur.execute(
                        "SELECT DISTINCT tenant_key FROM nb_session_messages ORDER BY tenant_key LIMIT 200"
                    )
                    rows = await cur.fetchall()
            return JSONResponse({"tenants": [r[0] for r in rows if r[0]]})
        except Exception as exc:
            logger.warning("admin list_tenants: {}", exc)
            return JSONResponse({"tenants": []})

    @router.get("/tenant/docs")
    async def tenant_docs(tenant_key: str = ""):
        """返回租户的 USER.md 内容 + 记忆索引 + Journey 阶段。"""
        if not tenant_key:
            return JSONResponse({"error": "missing tenant_key"}, status_code=400)
        result: dict = {"tenant_key": tenant_key, "user_md": "", "soul_md": "", "memory_index": [], "journey_stage": "novice"}
        try:
            if document_repo is not None:
                user_md = await document_repo.get(tenant_key, "USER.md")
                result["user_md"] = user_md or ""
                soul_md = await document_repo.get(tenant_key, "SOUL.md")
                result["soul_md"] = soul_md or ""
            if db is not None:
                from Mojing.storage.memory_repo import MySQLMemory
                memory = MySQLMemory(db=db, tenant_key=tenant_key, source="main")
                items = await memory.retrieve(top_k=20)
                result["memory_index"] = [
                    {"key": it.key, "description": it.description or it.content[:60]}
                    for it in items
                ]
            if tenant_state_repo is not None:
                stage = await tenant_state_repo.get_stage(tenant_key)
                result["journey_stage"] = stage
        except Exception as exc:
            logger.warning("admin tenant_docs: {}", exc)
            result["error"] = str(exc)
        return JSONResponse(result)

    @router.put("/tenant/doc")
    async def save_tenant_doc(request: Request):
        """保存租户的 USER.md 到数据库。"""
        try:
            body = await request.json()
        except Exception as exc:
            return Response(f"bad request: {exc}", status_code=400)
        tenant_key = str(body.get("tenant_key") or "").strip()
        doc_name = str(body.get("doc_name") or "").strip()
        content = str(body.get("content") or "")
        if not tenant_key or not doc_name:
            return Response("missing tenant_key or doc_name", status_code=400)
        if document_repo is None:
            return Response("document_repo not configured", status_code=503)
        try:
            await document_repo.set(tenant_key, doc_name, content)
            logger.info("admin: saved {}/{}", tenant_key, doc_name)
        except Exception as exc:
            return Response(str(exc), status_code=500)
        return Response(status_code=204)

    @router.get("/session/history")
    async def session_history(tenant_key: str = "", session_key: str = "", limit: int = 40):
        """返回会话历史消息（user + assistant 角色）。"""
        if not tenant_key or not session_key:
            return JSONResponse({"error": "missing tenant_key or session_key"}, status_code=400)
        if session_repo is None:
            return JSONResponse({"messages": []})
        try:
            stored, _ = await session_repo.load_messages(tenant_key, session_key)
        except Exception as exc:
            return JSONResponse({"messages": [], "error": str(exc)})
        msgs = []
        for m in stored:
            role = m.get("role", "")
            if role not in ("user", "assistant"):
                continue
            content = m.get("content", "") or ""
            if isinstance(content, list):
                content = " ".join(
                    p.get("text", "") for p in content
                    if isinstance(p, dict) and p.get("type") == "text"
                )
            if content:
                msgs.append({"role": role, "content": str(content)})
        return JSONResponse({"messages": msgs[-limit:]})

    @router.get("/tenant/dynamic_state")
    async def dynamic_state(tenant_key: str = "", topic_key: str = ""):
        """返回已停用的 topic reminder 调试状态。

        cold path 已改为 obligation ledger，不再向主 Agent 注入 topic reminder。
        该接口保留给旧调试页，避免调用方报错。
        """
        if not tenant_key:
            return JSONResponse({"error": "missing tenant_key"}, status_code=400)
        actual_topic_key = (topic_key or "").strip() or tenant_key
        result: dict = {
            "tenant_key": tenant_key,
            "topic_key":  actual_topic_key,
            "topic_state": None,
            "reminder_text": "",
            "disabled": True,
            "message": "topic reminder has been disabled; cold path now extracts obligations only.",
        }
        return JSONResponse(result)

    @router.get("/runtime/tasks")
    async def runtime_tasks(tenant_key: str = "", limit: int = 20):
        """返回最近的后台任务状态摘要。"""
        if runtime_task_repo is None:
            return JSONResponse({"tasks": [], "error": "runtime_task_repo not configured"})
        try:
            rows = await runtime_task_repo.list_recent(
                tenant_key=tenant_key.strip(),
                limit=limit,
            )
        except Exception as exc:
            logger.warning("admin runtime_tasks: {}", exc)
            return JSONResponse({"tasks": [], "error": str(exc)})
        return JSONResponse({
            "tenant_key": tenant_key.strip() or None,
            "tasks": rows,
        })

    # ================================================================
    # Prompt 预览（不调 LLM，只组装）
    # ================================================================

    @router.post("/preview/main_agent")
    async def preview_main_agent(request: Request):
        """预览主 Agent 运行时完整 Prompt：static + dynamic + tools。"""
        try:
            body = await request.json()
        except Exception:
            body = {}
        tenant_key = str(body.get("tenant_key") or "").strip()
        stage = str(body.get("stage") or "novice")
        try:
            from Mojing.config import load_stable_sections
            stable_parts = load_stable_sections(stage=stage)
            stable_prefix = "\n\n---\n\n".join(stable_parts)

            dynamic_parts: list[str] = []
            reminder_text = ""
            tool_schemas: list[dict] = []

            if tenant_key and main_agent is not None:
                dynamic_parts = []
                tool_schemas = main_agent.make_tool_registry(tenant_key, stage).schemas()
            elif tenant_key and document_repo is not None:
                # main_agent 未注入时的退化：至少把 USER.md 拿出来
                user_md = await document_repo.get(tenant_key, "USER.md")
                if user_md:
                    dynamic_parts = [user_md.strip()]

            full_prompt = _assemble_runtime_prompt(
                stable_prefix=stable_prefix,
                dynamic_parts=dynamic_parts,
                reminder=reminder_text,
                tool_schemas=tool_schemas,
            )
            return JSONResponse({
                "full_prompt":   full_prompt,
                "stage":         stage,
                "stable_parts":  len(stable_parts),
                "dynamic_parts": len(dynamic_parts),
                "has_reminder":  bool(reminder_text),
                "tools_count":   len(tool_schemas),
            })
        except Exception as exc:
            logger.exception("preview_main_agent")
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.post("/preview/skin_diary_agent")
    async def preview_skin_diary_agent(request: Request):
        """预览肌肤日记子 Agent 运行时完整 Prompt：static + dynamic + tools。"""
        try:
            body = await request.json()
        except Exception:
            body = {}
        tenant_key = str(body.get("tenant_key") or "").strip()
        try:
            stable_parts: list[str] = []
            dynamic_parts: list[str] = []
            reminder_text = ""
            tool_schemas: list[dict] = []

            if skin_diary_subagent is not None:
                # 子 Agent 的 stable section 就是 skin_diary.md（其 ContextBuilder 用 [self._prompt]）
                stable_parts = [skin_diary_subagent._prompt] if skin_diary_subagent._prompt else []
                if tenant_key:
                    dynamic_parts = []
                    tool_schemas = skin_diary_subagent.make_tool_registry(tenant_key).schemas()
            else:
                # 退化路径：从文件读 prompt
                skin_diary_entry = _prompt_map.get("skin_diary")
                if skin_diary_entry:
                    stable_parts = [skin_diary_entry.path.read_text(encoding="utf-8").strip()]

            stable_prefix = "\n\n---\n\n".join(stable_parts)
            full_prompt = _assemble_runtime_prompt(
                stable_prefix=stable_prefix,
                dynamic_parts=dynamic_parts,
                reminder=reminder_text,
                tool_schemas=tool_schemas,
            )
            return JSONResponse({
                "full_prompt":   full_prompt,
                "stable_parts":  len(stable_parts),
                "dynamic_parts": len(dynamic_parts),
                "has_reminder":  bool(reminder_text),
                "tools_count":   len(tool_schemas),
            })
        except Exception as exc:
            logger.exception("preview_skin_diary_agent")
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.post("/preview/deep_report_agent")
    async def preview_deep_report_agent(request: Request):
        """预览深度报告子 Agent 运行时完整 Prompt：static + dynamic + tools。

        可选 body.report_id 走 V2 三层 fallback（命中 → 注入对应报告；不命中 → latest）。
        """
        try:
            body = await request.json()
        except Exception:
            body = {}
        tenant_key = str(body.get("tenant_key") or "").strip()
        report_id = (str(body.get("report_id") or body.get("reportId") or "").strip()) or None
        try:
            stable_parts: list[str] = []
            dynamic_parts: list[str] = []
            reminder_text = ""
            tool_schemas: list[dict] = []

            if deep_report_subagent is not None:
                stable_parts = [deep_report_subagent._prompt] if deep_report_subagent._prompt else []
                if tenant_key:
                    dynamic_parts = []
                    tool_schemas = deep_report_subagent.make_tool_registry(tenant_key).schemas()
            else:
                deep_report_entry = _prompt_map.get("deep_report")
                if deep_report_entry:
                    stable_parts = [deep_report_entry.path.read_text(encoding="utf-8").strip()]

            stable_prefix = "\n\n---\n\n".join(stable_parts)
            full_prompt = _assemble_runtime_prompt(
                stable_prefix=stable_prefix,
                dynamic_parts=dynamic_parts,
                reminder=reminder_text,
                tool_schemas=tool_schemas,
            )
            return JSONResponse({
                "full_prompt":   full_prompt,
                "stable_parts":  len(stable_parts),
                "dynamic_parts": len(dynamic_parts),
                "has_reminder":  bool(reminder_text),
                "tools_count":   len(tool_schemas),
            })
        except Exception as exc:
            logger.exception("preview_deep_report_agent")
            return JSONResponse({"error": str(exc)}, status_code=500)

    # ================================================================
    # LLM 调用
    # ================================================================

    @router.post("/run/cold_path")
    async def run_cold_path(request: Request):
        """实际调用冷链路 LLM（单轮 obligation 抽取）。"""
        if llm is None:
            return JSONResponse({"error": "llm not configured"}, status_code=503)
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

        user_message = str(body.get("user_message") or "")
        assistant_reply = str(body.get("assistant_reply") or "")

        try:
            from Mojing.agent.cold_path import (
                _fill_user_template, _llm_complete_system_user, _parse_json_safe,
                _load_split_prompt, _COLD_PATH_PROMPT_PATH,
            )
            system_prompt, user_template = _load_split_prompt(_COLD_PATH_PROMPT_PATH)
            if not system_prompt or not user_template:
                return JSONResponse({"error": "cold_path.md missing or no SPLIT marker"}, status_code=500)

            user_content = _fill_user_template(
                template=user_template,
                user_message=user_message,
                assistant_reply=assistant_reply,
            )
            raw = await _llm_complete_system_user(
                llm, system=system_prompt, user=user_content, max_tokens=400,
            )
            parsed = _parse_json_safe(raw) if raw else None
            return JSONResponse({
                "raw": raw,
                "parsed": parsed,
                "system_len": len(system_prompt),
                "user_len": len(user_content),
            })
        except Exception as exc:
            logger.exception("admin run_cold_path")
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.post("/run/cold_path_chain")
    async def run_cold_path_chain(request: Request):
        """多轮顺序执行冷链路 obligation 抽取。"""
        if llm is None:
            return JSONResponse({"error": "llm not configured"}, status_code=503)
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

        rounds = body.get("rounds") or []
        if not isinstance(rounds, list) or not rounds:
            return JSONResponse({"error": "rounds must be a non-empty list"}, status_code=400)

        trace = []

        try:
            from Mojing.agent.cold_path import (
                _fill_user_template, _llm_complete_system_user, _parse_json_safe,
                _load_split_prompt, _COLD_PATH_PROMPT_PATH,
            )
            system_prompt, user_template = _load_split_prompt(_COLD_PATH_PROMPT_PATH)
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=500)

        for i, rd in enumerate(rounds):
            user_msg = str(rd.get("user_message") or "")
            asst_reply = str(rd.get("assistant_reply") or "")
            user_content = _fill_user_template(
                template=user_template,
                user_message=user_msg,
                assistant_reply=asst_reply,
            )
            raw = await _llm_complete_system_user(
                llm, system=system_prompt, user=user_content, max_tokens=400,
            )
            parsed = _parse_json_safe(raw) if raw else None
            trace.append({
                "round": i + 1,
                "user_message":  user_msg,
                "raw":    raw,
                "parsed": parsed,
            })

        return JSONResponse({"trace": trace})

    @router.post("/run/postprocess")
    async def run_postprocess(request: Request):
        """实际调用 Postprocess LLM（更新 USER.md）。"""
        if llm is None:
            return JSONResponse({"error": "llm not configured"}, status_code=503)
        try:
            body = await request.json()
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

        tenant_key = str(body.get("tenant_key") or "").strip()
        user_message = str(body.get("user_message") or "")
        assistant_reply = str(body.get("assistant_reply") or "")
        current_user_md = str(body.get("current_user_md") or "")

        # 从 DB 自动加载（如果前端没传）
        if tenant_key and not current_user_md and document_repo is not None:
            current_user_md = await document_repo.get(tenant_key, "USER.md") or ""

        try:
            from Mojing.agent.postprocess import (
                _render_input, _load_user_md_template, _load_postprocess_prompt,
                _UpdateDocTool,
            )
            from simpleclaw.core.loop import ReactLoop
            from simpleclaw.core.events import TextEvent, ToolResultEvent
            from simpleclaw.tools.base import ToolResult
            from simpleclaw.tools.registry import ToolRegistry

            system_prompt = _load_postprocess_prompt()
            user_md_template = _load_user_md_template()
            input_text = _render_input(
                user_message=user_message,
                assistant_reply=assistant_reply,
                user_md=current_user_md,
                user_md_template=user_md_template,
            )

            # Dry-run UpdateDoc：复用真实工具的 schema（让 LLM 参数名正确），
            # 但 execute 只记录调用，不写 DB。生产路径使用 _UpdateDocTool 真写。
            recorded_calls: list[dict] = []

            class _DryRunUpdateDocTool(_UpdateDocTool):
                def __init__(self) -> None:  # 绕过父类的 doc_repo 依赖
                    self._tenant_key = tenant_key
                    self.changed_docs = []

                async def execute(self, *, doc_key: str = "", content: str = "", **_) -> ToolResult:
                    recorded_calls.append({"doc_key": doc_key, "content": content})
                    return ToolResult(content=f"[dry-run] would update {doc_key}")

            registry = ToolRegistry()
            registry.register(_DryRunUpdateDocTool())

            loop = ReactLoop(llm=llm, tool_registry=registry, system_prompt=system_prompt, max_iterations=3)

            output_parts: list[str] = []
            async for event in loop.run(input_text):
                if isinstance(event, TextEvent):
                    output_parts.append(event.token)
                # ToolResultEvent 不展示（dry-run 结果无意义）

            # 把 LLM 实际收到的消息拼成一段可读文本，方便 admin 调试查看
            full_prompt = (
                "=== SYSTEM ===\n"
                + system_prompt
                + "\n\n=== USER ===\n"
                + input_text
            )

            return JSONResponse({
                "output": "".join(output_parts).strip(),
                "tool_calls": recorded_calls,       # ← 结构化展示 LLM 想写什么
                "full_prompt": full_prompt,         # ← system + user 完整 prompt
                "input_prompt": input_text,         # 保留供老 UI 兼容
                "system_prompt": system_prompt,
            })
        except Exception as exc:
            logger.exception("admin run_postprocess")
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.get("/tenant/skin_diary_result")
    async def tenant_skin_diary_result(tenant_key: str = "", diary_date: str = ""):
        """返回该租户肌肤日记分析结果，包含最新结果和指定业务日期历史。"""
        if not tenant_key:
            return JSONResponse({"error": "missing tenant_key"}, status_code=400)
        if db is None:
            return JSONResponse({"result": None})
        try:
            from Mojing.storage.skin_diary_result_repo import SkinDiaryResultRepository
            from Mojing.subagent.skin_diary import _format_analysis, _format_analysis_history
            result_repo = SkinDiaryResultRepository(db)
            latest = await result_repo.get_latest(tenant_key)
            if latest is None:
                return JSONResponse({"result": None})
            results = []
            if diary_date:
                from datetime import datetime
                parsed_date = datetime.strptime(diary_date, "%Y-%m-%d").date()
                results = await result_repo.get_results_for_business_date(tenant_key, parsed_date)
            elif latest.get("diary_date"):
                from datetime import datetime
                parsed_date = datetime.strptime(str(latest["diary_date"]), "%Y-%m-%d").date()
                results = await result_repo.get_results_for_business_date(tenant_key, parsed_date)
            return JSONResponse({
                "result": latest,
                "results": results,
                "formatted": _format_analysis(latest),
                "formatted_history": _format_analysis_history(results),
            })
        except Exception as exc:
            logger.warning("admin tenant_skin_diary_result: {}", exc)
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.get("/tenant/deep_report_list")
    async def tenant_deep_report_list(tenant_key: str = "", limit: int = 20):
        """返回该租户所有 status=done 的报告列表（admin 测试用，仅 slow 表元信息）。"""
        if not tenant_key:
            return JSONResponse({"error": "missing tenant_key"}, status_code=400)
        if db is None:
            return JSONResponse({"results": []})
        try:
            limit = max(1, min(int(limit), 100))
            sql = (
                "SELECT report_id, status, summary, create_time "
                "FROM nb_slow_model_reports "
                "WHERE user_id=%s AND deleted=0 AND status='done' "
                "ORDER BY create_time DESC LIMIT %s"
            )
            async with db.acquire() as conn:
                async with conn.cursor() as cur:
                    await cur.execute(sql, (tenant_key, limit))
                    rows = await cur.fetchall()
            results = [
                {
                    "report_id": r[0],
                    "status": r[1],
                    "summary": r[2],
                    "create_time": r[3].strftime("%Y-%m-%d %H:%M:%S") if r[3] else None,
                }
                for r in rows
            ]
            return JSONResponse({"results": results})
        except Exception as exc:
            logger.warning("admin tenant_deep_report_list: {}", exc)
            return JSONResponse({"error": str(exc)}, status_code=500)

    @router.get("/tenant/deep_report")
    async def tenant_deep_report(tenant_key: str = "", report_id: str = ""):
        """返回该租户的深度分析报告（V2 三表 JOIN）。

        - 传了 report_id：按 reportId 命中（双校验 user_id + report_id）；不命中也不 fallback，
          直接返回 None 让前端能区分"找不到"vs"未传"。
        - 没传 report_id：返回该用户最新一条 status=done 报告。
        """
        if not tenant_key:
            return JSONResponse({"error": "missing tenant_key"}, status_code=400)
        if db is None:
            return JSONResponse({"result": None})
        try:
            from Mojing.storage.deep_report_repo import DeepReportRepository
            from Mojing.subagent.deep_report import _format_deep_report
            repo = DeepReportRepository(db)
            rid = (report_id or "").strip()
            if rid:
                row = await repo.find_by_report_id_full(tenant_key, rid)
                source = f"by_report_id={rid}"
            else:
                row = await repo.find_latest_full(tenant_key)
                source = "latest"
            if row is None:
                return JSONResponse({"result": None, "source": source})
            formatted = ""
            try:
                formatted = _format_deep_report(row)
            except Exception as exc:
                logger.warning("admin _format_deep_report failed: {}", exc)

            return JSONResponse({
                "result": row,
                "formatted": formatted,
                "source": source,
            })
        except Exception as exc:
            logger.warning("admin tenant_deep_report: {}", exc)
            return JSONResponse({"error": str(exc)}, status_code=500)

    return router
