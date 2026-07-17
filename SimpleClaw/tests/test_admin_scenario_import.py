from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook

from admin.routes import _parse_xlsx_turns, _scenario_suite_from_admin_payload


def test_parse_xlsx_turns_extracts_user_time_and_media_url() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "samples"
    sheet.append(["发生时间", "用户发送的消息", "图片URL", "备注"])
    sheet.append([
        datetime(2026, 6, 10, 0, 59, 24),
        "最近618我想换防晒，帮我看看这一款适不适合我",
        "https://example.test/product.jpg",
        "",
    ])
    sheet.append(["", "", "", ""])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert payload["sheet"] == "samples"
    assert payload["turns"] == [
        {
            "row": 2,
            "occurred_at": "2026-06-10 00:59:24",
            "user": "最近618我想换防晒，帮我看看这一款适不适合我",
            "media": ["https://example.test/product.jpg"],
            "wait_after_s": 0,
            "baseline_reply": "",
            "review_mode": "stable",
            "product_feedback": "",
        }
    ]


def test_parse_xlsx_turns_keeps_user_rows_and_skips_assistant_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "6.10测试对话"
    sheet.append(["发送方", "发生时间", "内容", "图片链接", "等待秒"])
    sheet.append(["用户", "2026/6/10 0:59:24", "帮我看看这张脸", "https://example.test/face.jpg", 2])
    sheet.append(["assistant", "2026/6/10 1:00:00", "可以，我先看一下", "", ""])
    sheet.append(["客户", "2026/6/10 1:01:00", "那防晒呢 https://example.test/sunscreen.jpg", "", "5秒"])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert payload["sheet"] == "6.10测试对话"
    assert [turn["user"] for turn in payload["turns"]] == ["帮我看看这张脸", "那防晒呢"]
    assert payload["turns"][0]["media"] == ["https://example.test/face.jpg"]
    assert payload["turns"][1]["media"] == ["https://example.test/sunscreen.jpg"]
    assert [turn["wait_after_s"] for turn in payload["turns"]] == [2, 5]
    assert payload["skipped_rows"][0]["reason"] == "assistant_row"


def test_parse_xlsx_turns_maps_score_annotation_improve() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["发生时间", "用户消息", "评分标注", "产品反馈"])
    sheet.append(["2026-06-10 01:00:00", "帮我看看", "提升", "这里需要更保守"])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert payload["turns"][0]["review_mode"] == "improve"
    assert payload["turns"][0]["product_feedback"] == "这里需要更保守"


def test_parse_xlsx_turns_attaches_assistant_chat_annotation_to_previous_user() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["time", "role", "message_type", "content", "image_url", "评分标注", "反馈"])
    sheet.append(["2026-06-10 01:00:00", "user", "text", "我今天要外出", "", "", ""])
    sheet.append([
        "2026-06-10 01:00:05",
        "assistant",
        "text",
        "[CHAT] 知道啦，我帮你梳理下哦",
        "",
        "提升",
        "可以引导用户生成肌肤日记",
    ])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert payload["turns"][0]["baseline_reply"] == "知道啦，我帮你梳理下哦"
    assert payload["turns"][0]["review_mode"] == "improve"
    assert payload["turns"][0]["product_feedback"] == "可以引导用户生成肌肤日记"
    assert payload["review_counts"]["improve"] == 1
    assert payload["score_ready_count"] == 1


def test_parse_xlsx_turns_falls_back_to_best_text_cell() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["编号", "记录时间", "原始记录", "备注"])
    sheet.append([1, "2026-06-10 00:59:24", "用户问：我今天脸上好像有点痘痘", "首轮"])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert len(payload["turns"]) == 1
    assert payload["turns"][0]["user"] == "用户问：我今天脸上好像有点痘痘"
    assert payload["raw_rows"][0]["row"] == 2


def test_parse_xlsx_turns_handles_product_test_export_shape() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "6.10测试对话"
    sheet.append([
        "time", "role", "message_type", "content", "image_url", "voice_url",
        "voice_text", "tool_name", "tool_call_id", "tool_payload", "raw", "标注类型", "产品反馈",
    ])
    image_url = "https://example.test/product.jpg"
    sheet.append([
        "2026-06-10 00:59:24",
        "user",
        "image",
        image_url,
        image_url,
        "",
        "",
        "",
        "",
        "",
        '{"sender_type":"user","message_type":"image","content":"最近618我想换防晒","image_url":"https://example.test/product.jpg"}',
        "提升",
        "产品反馈：这里应该更保守一些",
    ])
    sheet.append([
        "2026-06-10 00:59:33",
        "user",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        '{"tenant_key":"395","session_key":"main:session_395","seq":59,"message_json":"{\\"role\\":\\"user\\",\\"content\\":\\"最近618我想换防晒\\"}"}',
        "",
        "",
    ])
    sheet.append(["2026-06-10 00:59:34", "assistant", "text", "收到啦", "", "", "", "", "", "", "", "", ""])
    sheet.append(["2026-06-10 01:00:00", "tool", "", "", "", "", "", "", "", '{"ok":true}', "", "", ""])
    sheet.append(["2026-06-10 01:01:00", "user", "image", "https://example.test/face.jpg", "https://example.test/face.jpg", "", "", "", "", "", '{"sender_type":"user","message_type":"image","content":"","image_url":"https://example.test/face.jpg"}', "不计分", ""])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = _parse_xlsx_turns(buffer.getvalue())

    assert payload["row_count"] == 5
    assert payload["turns"] == [
        {
            "row": 2,
            "occurred_at": "2026-06-10 00:59:24",
            "user": "",
            "media": [image_url],
            "wait_after_s": 0.0,
            "baseline_reply": "收到啦",
            "review_mode": "improve",
            "product_feedback": "产品反馈：这里应该更保守一些",
        },
        {
            "row": 6,
            "occurred_at": "2026-06-10 01:01:00",
            "user": "",
            "media": ["https://example.test/face.jpg"],
            "wait_after_s": 0.0,
            "baseline_reply": "",
            "review_mode": "none",
            "product_feedback": "",
        },
    ]
    assert [row["reason"] for row in payload["skipped_rows"][:3]] == [
        "persisted_duplicate_user_row",
        "assistant_row",
        "non_user_role:tool",
    ]


def test_scenario_suite_payload_supports_v1_device_entry() -> None:
    payload = _scenario_suite_from_admin_payload({
        "tenant_key": "test_device",
        "replay_surface": "v1_device",
        "device_id": "dev_001",
        "device_code": "mirror_001",
        "windows": {
            "main": {
                "turns": [{"user": "你好"}],
            },
        },
    })

    assert payload["replay"] == {
        "replay_surface": "v1_device",
        "protocol": "v1_chat_completions",
        "endpoint": "/v1/chat/completions",
        "prompt_surface": "device",
        "device_id": "dev_001",
        "device_code": "mirror_001",
        "custom": {
            "device_id": "dev_001",
            "device_code": "mirror_001",
        },
    }
